from flask import Blueprint, render_template, request, jsonify, current_app
from backend.utils.db_utils import add_import_history, get_import_history, clean_column_name
from backend.services.vanna_service import vanna_manager
import logging
import json
import pandas as pd
import numpy as np
import io
import os
import tempfile
from json import JSONEncoder  # 从标准库导入JSONEncoder
from backend.services.model_service import model_manager
from sqlalchemy import create_engine, text, inspect, types, MetaData # 确保导入 inspect 和 MetaData
from sqlalchemy.exc import SQLAlchemyError # 导入 SQLAlchemyError
from datetime import datetime
from decimal import Decimal
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

# 自定义 JSON 编码器处理 NumPy 和 Pandas 类型
class CustomJSONEncoder(JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif pd.isna(obj):
            return None
        elif hasattr(obj, 'item'): # 处理 numpy 标量
            return obj.item()
        elif isinstance(obj, (datetime.date, datetime.datetime)): # 处理日期时间
            return obj.isoformat()
        elif isinstance(obj, datetime.time): # 处理时间
            return obj.isoformat()
        elif isinstance(obj, Decimal): # 处理 Decimal
             return float(obj) # 或者 str(obj) 保持精度
        else:
            return super(CustomJSONEncoder, self).default(obj)

# 添加辅助函数，确保Pandas数据可以被JSON序列化
def pandas_to_json_safe(obj):
    """
    将Pandas/NumPy数据类型转换为JSON安全的Python原生类型
    """
    if isinstance(obj, pd.DataFrame):
        # 对DataFrame的每一列应用转换，特别是日期时间等
        # return obj.applymap(pandas_to_json_safe).to_dict(orient='records')
        # 使用 to_json 可能更方便处理日期等类型，然后再加载回python对象
        json_str = obj.to_json(orient='records', date_format='iso', default_handler=str) 
        return json.loads(json_str)
    elif isinstance(obj, pd.Series):
        # return obj.apply(pandas_to_json_safe).to_dict()
        json_str = obj.to_json(orient='records', date_format='iso', default_handler=str)
        return json.loads(json_str)
    elif pd.isna(obj):  # 处理NaN/None
        return None
    elif hasattr(obj, 'item'):  # 处理NumPy标量类型如np.int64等
        return obj.item()
    elif hasattr(obj, 'tolist'):  # 处理NumPy数组
        return obj.tolist()
    elif isinstance(obj, (datetime.date, datetime.datetime)): # 处理日期时间
            return obj.isoformat()
    elif isinstance(obj, datetime.time): # 处理时间
        return obj.isoformat()
    elif isinstance(obj, Decimal): # 处理 Decimal
        return float(obj)
    return obj

# 修改检测表标题的逻辑，扩展检测范围到前几行
def detect_and_process_excel_structure(file_obj, auto_detect_row_count=True, user_specified_row_count=1, title_row_fixed=None, header_start_row_fixed=None):
    """
    检测并处理Excel文件的结构，包括合并单元格、多级表头等
    
    参数:
    file_obj: 文件对象
    auto_detect_row_count: 是否自动检测表头行数
    user_specified_row_count: 用户指定的表头行数 (仅在 auto_detect_row_count=False 时优先使用)
    title_row_fixed: 固定的表标题行号 (从1开始，0或None表示自动检测)
    header_start_row_fixed: 固定的表头开始行号 (从1开始，0或None表示自动检测)
    
    返回:
    dict: 包含处理结果的字典
    """
    # 保存临时文件以便openpyxl打开
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp_path = tmp.name
        file_obj.save(tmp_path)
    
    try:
        # 使用openpyxl打开文件以检查合并单元格
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        sheet = wb.active
        
        # 获取合并单元格信息
        merged_cells = sheet.merged_cells.ranges
        merged_cell_info = []
        for merged_range in merged_cells:
            merged_cell_info.append({
                'min_row': merged_range.min_row,
                'max_row': merged_range.max_row,
                'min_col': merged_range.min_col,
                'max_col': merged_range.max_col,
                'value': sheet.cell(merged_range.min_row, merged_range.min_col).value
            })
        
        # --- 1. Determine Title Row --- 
        has_title_row = False
        title_text = None
        title_row = 0 # 0 means no title row found or specified as such
        
        if title_row_fixed is not None and title_row_fixed > 0:
            has_title_row = True
            title_row = title_row_fixed
            # Try to get value, considering merges
            title_text = sheet.cell(title_row, 1).value # Default value
            for mc in merged_cell_info:
                if mc['min_row'] == title_row and mc['min_col'] == 1:
                    title_text = mc['value'] # Prefer merged cell value
                    break 
        elif title_row_fixed is None or title_row_fixed == 0: # Auto-detect only if not explicitly set to non-zero
            # (Current auto-detection logic for title row)
            for r in range(1, 4): # Check first 3 rows
                # Check for wide merged cell
                wide_merged = [mc for mc in merged_cell_info if mc['min_row'] == r and 
                               (mc['max_col'] - mc['min_col'] > 2 or mc['max_col'] - mc['min_col'] >= sheet.max_column -2)]
                if wide_merged:
                    has_title_row, title_text, title_row = True, wide_merged[0]['value'], r
                    break
                # Check for single cell title pattern
                if not has_title_row and sheet.cell(r, 1).value and not sheet.cell(r,2).value and not sheet.cell(r,3).value:
                    has_title_row, title_text, title_row = True, sheet.cell(r,1).value, r
                    break
            if not has_title_row: # One last check for a single cell in the entire row (first few cols)
                for r_check_title in range(1,4):
                    row_values_title = [sheet.cell(r_check_title,c_title).value for c_title in range(1, min(sheet.max_column + 1, 6))]
                    non_empty_values_title = [v_title for v_title in row_values_title if v_title is not None and str(v_title).strip() != '']
                    if len(non_empty_values_title) == 1:
                        has_title_row, title_text, title_row = True, non_empty_values_title[0], r_check_title
                        break

        # --- 2. Determine Header Start Row --- 
        header_start_row = title_row + 1 if has_title_row else 1
        if header_start_row_fixed is not None and header_start_row_fixed > 0:
            header_start_row = header_start_row_fixed
        
        # Ensure header_start_row is at least 1
        if header_start_row <= 0 : header_start_row = 1

        # --- 3. Detect Header Complexity (local multi_level_header) --- 
        header_complex = False
        multi_level_header = False # This is locally detected, not from form
        # header_merged_cells are those within the potential header area (e.g. header_start_row to header_start_row + 4)
        header_merged_cells = [mc for mc in merged_cell_info if 
                               mc['min_row'] >= header_start_row and 
                               mc['min_row'] < header_start_row + 5] # Check up to 5 rows for header merges
        if header_merged_cells:
            header_complex = True
            if any(mc['max_col'] > mc['min_col'] for mc in header_merged_cells) or \
               any(mc['max_row'] > mc['min_row'] for mc in header_merged_cells):
                multi_level_header = True

        # --- 4. Determine Actual Header Rows (Count) --- 
        actual_header_rows = 1 # Default
        if not auto_detect_row_count: # User specified header_start_row, so their count is authority
            actual_header_rows = user_specified_row_count
        else:
            # Auto-detect the number of header rows
            _detected_count = 1
            if multi_level_header: # Use local detection of multi_level_header
                max_depth_found = header_start_row - 1
                # Consider merges that start within the determined header_start_row and span downwards
                header_area_merges = [mc for mc in merged_cell_info if 
                                      mc['min_row'] >= header_start_row and 
                                      mc['min_row'] < header_start_row + 5] # Look for merges within 5 rows from start
                if header_area_merges:
                    # The deepest point a merge reaches from the header_start_row downwards
                    max_depth_found = max(mc['max_row'] for mc in header_area_merges)
                
                if max_depth_found >= header_start_row:
                    _detected_count = max_depth_found - header_start_row + 1
                else: # No defining merges, check continuous non-empty rows from header_start_row
                    _cont_rows = 0
                    for i in range(5): # Check up to 5 rows from header_start_row
                        r_check = header_start_row + i
                        if r_check > sheet.max_row: break # Past end of sheet
                        # Check if row has any content in first few columns
                        if not any(sheet.cell(r_check, c).value for c in range(1, min(sheet.max_column + 1, 10))):
                            break # Empty row marks end of header
                        _cont_rows += 1
                    _detected_count = _cont_rows if _cont_rows > 0 else 1
            else: # Simple header (no multi-level characteristics detected)
                 _detected_count = 1 # Usually 1 row for simple headers
            
            actual_header_rows = min(_detected_count, 5) # Cap auto-detected rows to 5
        
        if actual_header_rows <= 0: actual_header_rows = 1 # Ensure positive count

        # 构建正确的多行表头
        header_values = []
        # Initially read up to sheet.max_column as a potential maximum
        initial_max_col_read = sheet.max_column 
        for r_idx in range(header_start_row, header_start_row + actual_header_rows):
            if r_idx > sheet.max_row: break 
            row_vals = []
            for c_idx in range(1, initial_max_col_read + 1):
                is_merged_cell = False
                cell_val = None
                for mc_range in merged_cell_info:
                    if mc_range['min_row'] <= r_idx <= mc_range['max_row'] and \
                       mc_range['min_col'] <= c_idx <= mc_range['max_col']:
                        cell_val = mc_range['value'] 
                        is_merged_cell = True
                        break
                if not is_merged_cell:
                    cell_val = sheet.cell(r_idx, c_idx).value
                row_vals.append(cell_val)
            header_values.append(row_vals)

        # Determine the actual maximum column with data in the headers
        final_max_col = 0
        if header_values and header_values[0]: # Ensure header_values is not empty and has at least one row
            # Iterate from right to left over columns of the first header row (or longest if they vary)
            # Assume all header rows should have same number of columns after initial read
            num_cols_in_header_row = len(header_values[0])
            for c_idx_from_zero in range(num_cols_in_header_row - 1, -1, -1):
                column_has_content = False
                for r_values in header_values: # Check this column index in all header rows
                    if c_idx_from_zero < len(r_values): # Check if this row has this column
                        if r_values[c_idx_from_zero] is not None and str(r_values[c_idx_from_zero]).strip() != "":
                            column_has_content = True
                            break # Found content in this column in one of the header rows
                if column_has_content:
                    final_max_col = c_idx_from_zero + 1 # Convert 0-based index to 1-based count
                    break # Found the rightmost column with content
        
        if final_max_col == 0: # If no content found in any header cell or headers are empty
            if initial_max_col_read > 0: # If sheet had columns but headers were blank
                final_max_col = 1 # Default to at least one column if sheet.max_column was > 0
            else: # sheet.max_column was 0 or headers were non-existent
                 final_max_col = 0

        # Trim header_values to the final_max_col if it was reduced
        if final_max_col > 0 and final_max_col < initial_max_col_read:
            trimmed_header_values = []
            for r_vals in header_values:
                trimmed_header_values.append(r_vals[:final_max_col])
            header_values = trimmed_header_values
        elif final_max_col == 0 and header_values: # All headers were blank, effectively no header data
            header_values = [[] for _ in header_values] # Make all header rows empty lists
        
        # 检测正文中的合并单元格
        body_start_row = header_start_row + actual_header_rows
        body_merged_cells = [mc for mc in merged_cell_info if mc['min_row'] >= body_start_row]
        
        # 按列分组的合并单元格 (通常是同一列内的相同值被合并)
        column_merged_cells = {}
        for mc in body_merged_cells:
            if mc['min_col'] == mc['max_col']:  # 仅限垂直合并的单元格
                col = mc['min_col']
                if col not in column_merged_cells:
                    column_merged_cells[col] = []
                column_merged_cells[col].append(mc)
        
        # 检测每列是否有垂直合并单元格
        columns_with_merged_cells = list(column_merged_cells.keys())
        
        # 读取有效数据的范围
        max_row_data = sheet.max_row
        has_merged_cells_in_body = len(body_merged_cells) > 0
        
        return {
            'has_title_row': has_title_row,
            'title_text': title_text,
            'title_row': title_row,
            'header_start_row': header_start_row,
            'header_complex': header_complex,
            'multi_level_header': multi_level_header, # Locally detected multi-level nature
            'actual_header_rows': actual_header_rows, # The count of header rows used
            'header_values': header_values, # Potentially trimmed header values
            'body_start_row': body_start_row,
            'max_row': max_row_data,
            'max_col': final_max_col, # Use the refined max_col
            'merged_cell_info': merged_cell_info,
            'has_merged_cells_in_body': has_merged_cells_in_body,
            'columns_with_merged_cells': columns_with_merged_cells,
            'column_merged_cells': column_merged_cells,
            'temp_file_path': tmp_path
        }
    except Exception as e:
        # 确保临时文件被删除
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise e

def create_dataframe_from_excel_structure(excel_info):
    """
    根据Excel结构信息创建Pandas DataFrame
    
    参数:
    excel_info: 来自detect_and_process_excel_structure的结构信息
    
    返回:
    pandas.DataFrame: 处理后的数据帧
    """
    try:
        # 打开临时文件
        wb = openpyxl.load_workbook(excel_info['temp_file_path'], data_only=True)
        sheet = wb.active
        
        # 确定表头和数据
        header_start = excel_info['header_start_row']
        header_rows = excel_info['actual_header_rows']
        data_start = excel_info['body_start_row']
        max_row = excel_info['max_row']
        max_col = excel_info['max_col']
        
        # 准备列名
        if excel_info['multi_level_header']:
            # 处理多级表头
            header_values = excel_info['header_values']
            
            # 构建多层次列名，避免重复
            if len(header_values) > 1:
                # 智能合并多行表头，避免冗余
                merged_headers = []
                for col_idx in range(max_col):
                    if col_idx >= len(header_values[0]):
                        merged_headers.append(f'Column_{col_idx+1}')
                        continue
                        
                    # 获取这一列的所有表头值
                    col_values = []
                    last_value = None
                    for row_idx in range(len(header_values)):
                        if col_idx < len(header_values[row_idx]):
                            value = header_values[row_idx][col_idx]
                            # 如果值与上一个不同且不为空，则添加
                            if value is not None and value != '' and value != last_value:
                                col_values.append(str(value).strip())
                                last_value = value
                    
                    # 如果没有有效值，使用默认列名
                    if not col_values:
                        merged_headers.append(f'Column_{col_idx+1}')
                    else:
                        # 合并非重复值
                        merged_headers.append('_'.join(col_values).replace('\n', ''))
                
                columns = merged_headers
            else:
                columns = [str(val).strip().replace('\n', '') if val is not None else f'Column_{i+1}' 
                          for i, val in enumerate(header_values[0])]
        else:
            # 单行表头
            columns = []
            for col in range(1, max_col + 1):
                val = sheet.cell(header_start, col).value
                columns.append(str(val).strip().replace('\n', '') if val is not None else f'Column_{col}')
        
        # 准备数据
        data = []
        for row in range(data_start, max_row + 1):
            row_data = []
            for col in range(1, max_col + 1):
                # 检查当前单元格是否在合并单元格内
                cell_value = None
                for mc in excel_info['merged_cell_info']:
                    if (mc['min_row'] <= row <= mc['max_row'] and 
                        mc['min_col'] <= col <= mc['max_col']):
                        # 使用合并单元格的值
                        cell_value = mc['value']
                        break
                
                if cell_value is None:
                    cell_value = sheet.cell(row, col).value
                
                row_data.append(cell_value)
            
            if any(val is not None for val in row_data):  # 跳过空行
                data.append(row_data)
        
        # 创建DataFrame
        df = pd.DataFrame(data, columns=columns)
        
        # 处理垂直方向的合并单元格 (向下填充)
        for col_idx, merged_cells_list in excel_info['column_merged_cells'].items():
            if col_idx <= len(columns):
                col_name = columns[col_idx - 1]
                # 处理NaN值后执行前向填充
                df[col_name] = df[col_name].fillna(method='ffill')
        
        # 清理数据 - 删除全为空的行
        df = df.dropna(how='all')
        
        # 如果有标题行，添加为元数据而不是列前缀
        if excel_info['has_title_row'] and excel_info['title_text']:
            # 不要将标题添加到每个列名前
            df.attrs['title'] = excel_info['title_text']
        
        return df
    
    finally:
        # 删除临时文件
        if os.path.exists(excel_info['temp_file_path']):
            os.unlink(excel_info['temp_file_path'])

data_import_bp = Blueprint('data_import', __name__)
logger = logging.getLogger(__name__)
app = current_app



@data_import_bp.route('/data-import')
def data_import():
    return render_template('data_import.html')

@data_import_bp.route('/api/upload-file', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({"status": "error", "message": "没有上传文件"}), 400
            
        file = request.files['file']
        if file.filename == '':
            return jsonify({"status": "error", "message": "未选择文件"}), 400
            
        # 获取文件名和扩展名
        filename = file.filename
        file_ext = filename.rsplit('.', 1)[1].lower()
        
        # 获取导入选项参数
        # has_header is determined by 'header' checkbox (value 'on' if checked) OR by explicit 'has_header' form data
        # If 'header' checkbox is present and checked, request.form.get('header') will be 'on'.
        # If frontend explicitly sets formData.append('has_header', 'false'), then request.form.get('has_header') is 'false'.
        # We prioritize an explicit 'has_header' field from FormData if present.
        has_header_explicit = request.form.get('has_header') # This would be 'true' or 'false' if set by JS directly
        
        if has_header_explicit is not None:
            has_header = has_header_explicit == 'true'
        else:
            # Fallback to checking the 'header' checkbox if 'has_header' was not explicitly set in FormData
            has_header = request.form.get('header') == 'on' # 'on' is the typical value for a checked checkbox
            
        custom_headers = json.loads(request.form.get('custom_headers', '[]'))
        process_complex_structure = request.form.get('process_complex_structure', 'true') == 'true'
        detect_title_row = request.form.get('detect_title_row', 'true') == 'true'
        multi_level_header = request.form.get('multi_level_header', 'auto')  # auto, true, false
        fill_merged_cells = request.form.get('fill_merged_cells', 'true') == 'true'
        
        # 获取用户指定的结构范围参数
        title_row = int(request.form.get('title_row', '0'))  # 0表示无标题
        header_start_row = int(request.form.get('header_start_row', '0'))  # 0表示自动检测
        header_rows = int(request.form.get('header_rows', '1'))
        
        # 根据文件类型使用pandas读取
        df = None
        structure_info = None
        
        if file_ext == 'csv':
            # CSV文件处理 - 不包含合并单元格，但需要处理表头
            if has_header:
                df = pd.read_csv(file)
            else:
                if custom_headers:
                    df = pd.read_csv(file, header=None, names=custom_headers)
                else:
                    df = pd.read_csv(file, header=None)
                    df.columns = [f'Column_{i+1}' for i in range(len(df.columns))]
        elif file_ext in ['xlsx', 'xls']:
            # Excel文件处理 - 可能包含复杂结构
            if process_complex_structure:
                # 处理合并单元格和复杂表结构
                try:
                    file.seek(0)  # 确保从文件开头读取
                    # Determine if user wants auto-detection for header configuration
                    auto_detect_header_config = (header_start_row == 0) # If start is auto, count is also effectively auto unless specified
                    
                    structure_info = detect_and_process_excel_structure(
                        file, 
                        auto_detect_row_count=auto_detect_header_config,
                        user_specified_row_count=header_rows, # This is the header_rows from form
                        title_row_fixed=title_row,
                        header_start_row_fixed=header_start_row
                    )
                    
                    # 重新打开文件
                    file.seek(0)
                    df = create_dataframe_from_excel_structure(structure_info)
                    
                    # 添加结构信息到预览中
                    structure_summary = {
                        'has_title': structure_info['has_title_row'],
                        'title_text': structure_info['title_text'],
                        'title_row': structure_info['title_row'],
                        'has_complex_header': structure_info['header_complex'],
                        'multi_level_header': structure_info['multi_level_header'],
                        'header_rows': structure_info['actual_header_rows'],
                        'header_start_row': structure_info['header_start_row'],
                        'has_merged_cells_in_body': structure_info['has_merged_cells_in_body'],
                        'columns_with_merged_cells': [int(col) for col in structure_info['columns_with_merged_cells']]
                    }
                except Exception as e:
                    logger.error(f"处理复杂Excel结构失败: {str(e)}")
                    # 如果复杂处理失败，回退到常规pandas处理
                    file.seek(0)
                    structure_info = None
                    # 继续使用下面的常规Excel处理
            
            # 如果未能通过复杂处理获取数据，使用pandas常规处理
            if df is None:
                if has_header:
                    # 使用pandas标准读取，根据用户选择的表头行数
                    header_param = 0 if header_rows == 1 else list(range(header_rows))
                    df = pd.read_excel(file, header=header_param)
                    
                    # 处理多级表头
                    if header_rows > 1:
                        df.columns = ['_'.join([str(x) if pd.notna(x) else '' for x in col]).strip('_') 
                                     for col in df.columns.values]
                else:
                    if custom_headers:
                        df = pd.read_excel(file, header=None, names=custom_headers)
                    else:
                        df = pd.read_excel(file, header=None)
                        df.columns = [f'Column_{i+1}' for i in range(len(df.columns))]
        else:
            return jsonify({"status": "error", "message": "不支持的文件格式"}), 400
            
        # 获取原始列名
        original_columns = df.columns.tolist()
        logger.info(f"原始列名: {original_columns}")
        # 准备样本数据
        sample_data = {}
        for col in original_columns:
            sample_data[col] = df[col].head(1).astype(str).tolist()
        
        # 检查是否使用AI分析
        use_ai = request.form.get('use_ai') == 'true'
        if use_ai:
            try:
                analysis = model_manager.analyze_columns(
                    columns=original_columns,
                    sample_data=sample_data,
                    total_rows=len(df)
                )
                
                # 构建预览数据
                preview_data = {
                    "columns": original_columns,
                    "data_types": pandas_to_json_safe(df.dtypes.astype(str)),
                    "preview": pandas_to_json_safe(df.head(5)),
                    "all_data": pandas_to_json_safe(df),
                    "total_rows": analysis.total_rows,
                    "total_columns": analysis.total_columns,
                    "column_analysis": [col.model_dump() for col in analysis.columns],
                    "file_name": filename,
                    "file_path": filename,  # 前端只能获取文件名
                    "import_options": {
                        "has_header": has_header,
                        "process_complex_structure": process_complex_structure,
                        "detect_title_row": detect_title_row,
                        "multi_level_header": multi_level_header,
                        "fill_merged_cells": fill_merged_cells,
                        "header_rows": header_rows
                    }
                }
                
                # 如果有Excel结构信息，添加到预览数据
                if structure_info:
                    preview_data["excel_structure"] = structure_summary
                
            except Exception as e:
                app.logger.error(f"列分析失败: {str(e)}")
                return jsonify({"status": "error", "message": f"列分析失败: {str(e)}"}), 500
        else:
            # 不使用AI分析时，提供基本的预览数据
            preview_data = {
                "columns": original_columns,
                "data_types": pandas_to_json_safe(df.dtypes.astype(str)),
                "preview": pandas_to_json_safe(df.head(5)),
                "all_data": pandas_to_json_safe(df),
                "total_rows": len(df),
                "total_columns": len(df.columns),
                "file_name": filename,
                "file_path": filename,  # 前端只能获取文件名
                "import_options": {
                    "has_header": has_header,
                    "process_complex_structure": process_complex_structure,
                    "detect_title_row": detect_title_row,
                    "multi_level_header": multi_level_header,
                    "fill_merged_cells": fill_merged_cells,
                    "header_rows": header_rows
                }
            }
            
            # 如果有Excel结构信息，添加到预览数据
            if structure_info:
                preview_data["excel_structure"] = structure_summary
        
        return jsonify({
            "status": "success",
            "message": "文件上传成功",
            "preview": preview_data
        })
        
    except Exception as e:
        app.logger.error(f"文件上传失败: {str(e)}")
        return jsonify({"status": "error", "message": f"文件上传失败: {str(e)}"}), 500

@data_import_bp.route('/api/existing-tables', methods=['GET'])
def get_existing_tables():
    """获取数据库中所有用户创建的表的列表"""
    try:
        config = vanna_manager.get_config()
        if not config or 'database' not in config:
            return jsonify({"status": "error", "message": "未配置数据库连接"}), 400

        db_config = config['database']
        engine = None
        try:
            if db_config['type'] == 'mysql':
                mysql_url = f"mysql+pymysql://{db_config['username']}:{db_config['password']}@{db_config['host']}:{db_config.get('port', 3306)}/{db_config['database_name']}?charset=utf8mb4"
                engine = create_engine(mysql_url)
            elif db_config['type'] == 'postgres':
                pg_url = f"postgresql+psycopg2://{db_config['username']}:{db_config['password']}@{db_config['host']}:{db_config.get('port', 5432)}/{db_config['database_name']}"
                engine = create_engine(pg_url)
            elif db_config['type'] == 'sqlite':
                sqlite_path = db_config['database_name']
                if not sqlite_path:
                     return jsonify({"status": "error", "message": "SQLite 数据库路径未配置"}), 400
                engine = create_engine(f'sqlite:///{sqlite_path}')
            else:
                 return jsonify({"status": "error", "message": "不支持的数据库类型"}), 400

            inspector = inspect(engine)
            # 获取所有表名，可以根据需要过滤系统表
            # 对于 MySQL/PostgreSQL，这通常会包含所有表
            # 对于 SQLite，这也会包含 sqlite_sequence 等内部表，可能需要过滤
            tables = inspector.get_table_names()
            # 可选：过滤掉特定的系统表（如果需要）
            # tables = [t for t in tables if not t.startswith('sqlite_')] # SQLite 示例

            return jsonify({"status": "success", "tables": tables})

        except Exception as e:
            logger.error(f"获取表列表失败: {str(e)}")
            return jsonify({"status": "error", "message": f"获取表列表失败: {str(e)}"}), 500
        finally:
            if engine:
                engine.dispose()

    except Exception as e:
        logger.error(f"获取表列表时发生错误: {str(e)}")
        return jsonify({"status": "error", "message": f"获取表列表时发生错误: {str(e)}"}), 500


@data_import_bp.route('/api/table-schema', methods=['GET'])
def get_table_schema():
    """获取指定表的结构（列名和数据类型）"""
    table_name = request.args.get('table_name')
    if not table_name:
        return jsonify({"status": "error", "message": "缺少表名参数"}), 400

    try:
        config = vanna_manager.get_config()
        if not config or 'database' not in config:
            return jsonify({"status": "error", "message": "未配置数据库连接"}), 400

        db_config = config['database']
        engine = None
        try:
            # 创建数据库连接（与上面类似）
            if db_config['type'] == 'mysql':
                mysql_url = f"mysql+pymysql://{db_config['username']}:{db_config['password']}@{db_config['host']}:{db_config.get('port', 3306)}/{db_config['database_name']}?charset=utf8mb4"
                engine = create_engine(mysql_url)
            elif db_config['type'] == 'postgres':
                 pg_url = f"postgresql+psycopg2://{db_config['username']}:{db_config['password']}@{db_config['host']}:{db_config.get('port', 5432)}/{db_config['database_name']}"
                 engine = create_engine(pg_url)
            elif db_config['type'] == 'sqlite':
                 sqlite_path = db_config['database_name']
                 if not sqlite_path:
                      return jsonify({"status": "error", "message": "SQLite 数据库路径未配置"}), 400
                 engine = create_engine(f'sqlite:///{sqlite_path}')
            else:
                  return jsonify({"status": "error", "message": "不支持的数据库类型"}), 400

            inspector = inspect(engine)
            if not inspector.has_table(table_name):
                 return jsonify({"status": "error", "message": f"表 '{table_name}' 不存在"}), 404

            columns = inspector.get_columns(table_name)
            schema = [{"name": col['name'], "type": str(col['type'])} for col in columns]

            return jsonify({"status": "success", "schema": schema})

        except Exception as e:
            logger.error(f"获取表 '{table_name}' 的结构失败: {str(e)}")
            return jsonify({"status": "error", "message": f"获取表 '{table_name}' 的结构失败: {str(e)}"}), 500
        finally:
            if engine:
                engine.dispose()

    except Exception as e:
        logger.error(f"获取表结构时发生错误: {str(e)}")
        return jsonify({"status": "error", "message": f"获取表结构时发生错误: {str(e)}"}), 500


@data_import_bp.route('/api/import-data', methods=['POST'])
def import_data():
    try:
        data = request.json
        if not data or 'data' not in data or 'table_name' not in data:
            return jsonify({"status": "error", "message": "缺少必要参数"}), 400

        import_mode = data.get('import_mode', 'replace') # 'replace' or 'append'
        table_name = data['table_name']
        df = pd.DataFrame(data['data'])
        column_settings = data.get('column_settings', {}) # { original_name: { type: '...' } }
        selected_columns = data.get('selected_columns', []) # [original_name, ...]
        # column_mapping: { original_name: target_name }
        # 对于 append 模式，target_name 是现有表的列名
        # 对于 replace 模式，target_name 是清理后的新列名
        column_mapping = data.get('column_mapping', {})
        primary_key_original = data.get('primary_key') # 仅用于 replace 模式
        # 获取Excel结构信息
        excel_structure = data.get('excel_structure', {})

        # --- 基本验证 ---
        if not selected_columns:
             return jsonify({"status": "error", "message": "没有选择任何列进行导入"}), 400

        missing_selected = [col for col in selected_columns if col not in df.columns]
        if missing_selected:
            return jsonify({
                "status": "error",
                "message": f"以下选中的列在原始数据中不存在: {', '.join(missing_selected)}",
            }), 400

        # 只保留选中的列
        df = df[selected_columns]
        
        # 处理Excel合并单元格中的空值（如果前端未处理）
        if excel_structure and excel_structure.get('has_merged_cells_in_body'):
            columns_with_merged_cells = excel_structure.get('columns_with_merged_cells', [])
            # 将列索引转换为列名
            for idx in columns_with_merged_cells:
                if 0 <= idx - 1 < len(selected_columns):
                    col_name = selected_columns[idx - 1]
                    # 前向填充空值
                    if col_name in df.columns:
                        df[col_name] = df[col_name].fillna(method='ffill')

        # 获取数据库配置和 engine
        config = vanna_manager.get_config()
        if not config or 'database' not in config:
            return jsonify({"status": "error", "message": "未配置数据库连接"}), 400
        db_config = config['database']
        engine = None

        try:
            # 创建 engine (复用上面的逻辑)
            if db_config['type'] == 'mysql':
                mysql_url = f"mysql+pymysql://{db_config['username']}:{db_config['password']}@{db_config['host']}:{db_config.get('port', 3306)}/{db_config['database_name']}?charset=utf8mb4"
                engine = create_engine(mysql_url)
            elif db_config['type'] == 'postgres':
                 pg_url = f"postgresql+psycopg2://{db_config['username']}:{db_config['password']}@{db_config['host']}:{db_config.get('port', 5432)}/{db_config['database_name']}"
                 engine = create_engine(pg_url)
            elif db_config['type'] == 'sqlite':
                 sqlite_path = db_config['database_name']
                 if not sqlite_path:
                     return jsonify({"status": "error", "message": "SQLite 数据库路径未配置"}), 400
                 engine = create_engine(f'sqlite:///{sqlite_path}')
            else:
                 return jsonify({"status": "error", "message": "不支持的数据库类型"}), 400

            # --- 根据导入模式处理 ---
            if import_mode == 'append':
                # --- 追加模式 ---
                inspector = inspect(engine)
                if not inspector.has_table(table_name):
                    return jsonify({"status": "error", "message": f"目标表 '{table_name}' 不存在，无法追加"}), 404

                # 获取目标表结构
                target_columns_info = inspector.get_columns(table_name)
                target_columns_map = {col['name']: str(col['type']) for col in target_columns_info}
                target_column_names = list(target_columns_map.keys())

                # 验证 column_mapping 中的目标列是否存在于表中
                final_mapping = {} # { original_file_col: target_table_col }
                mapped_target_cols = []
                for original_col, target_col in column_mapping.items():
                    if original_col in df.columns: # 确保原始列是被选中的
                        if target_col in target_column_names:
                             final_mapping[original_col] = target_col
                             mapped_target_cols.append(target_col)
                        elif target_col and target_col != 'do_not_import': # 忽略未映射或明确不导入的列
                            logger.warning(f"文件列 '{original_col}' 映射的目标表列 '{target_col}' 在表 '{table_name}' 中不存在，将被忽略。")


                # 筛选并重命名 DataFrame 列以匹配目标表
                cols_to_keep = list(final_mapping.keys())
                if not cols_to_keep:
                     return jsonify({"status": "error", "message": "没有有效的列映射到目标表"}), 400

                df_append = df[cols_to_keep].rename(columns=final_mapping)

                # 可选：根据目标表类型尝试转换数据类型 (更安全的方式是让数据库处理，但可以预处理)
                # 注意：这里简化处理，to_sql 通常能处理基本类型转换
                # for target_col in df_append.columns:
                #     target_type = target_columns_map.get(target_col)
                #     # ... (添加基于 target_type 的 pd.to_numeric, pd.to_datetime 等转换逻辑，注意错误处理)

                # 追加数据
                logger.info(f"开始向表 '{table_name}' 追加数据，列: {list(df_append.columns)}")
                df_append.to_sql(
                    name=table_name,
                    con=engine,
                    if_exists='append',
                    index=False,
                    method='multi' if db_config['type'] != 'sqlite' else None, # SQLite 不支持 multi
                    chunksize=1000 if db_config['type'] != 'sqlite' else None
                )
                rows_imported = len(df_append)
                logger.info(f"成功向表 '{table_name}' 追加 {rows_imported} 行数据")

                # 记录追加历史 (可以简化列信息，因为是追加)
                column_info_history = {
                    'mode': 'append',
                    'mapped_columns': final_mapping, # { original_file_col: target_table_col }
                    'target_columns': target_column_names,
                    'excel_structure': excel_structure if excel_structure else None
                }
                primary_key_final = None # 追加模式不设置主键

            elif import_mode == 'replace':
                # --- 替换模式 (大部分是现有逻辑) ---

                # 清理列名 (只对替换模式需要，用于创建新表)
                cleaned_column_mapping = {orig: clean_column_name(target) for orig, target in column_mapping.items() if orig in df.columns}
                df = df.rename(columns=cleaned_column_mapping)

                # 处理多级表头的列名清理
                if excel_structure and excel_structure.get('multi_level_header'):
                    for col in df.columns:
                        if col in cleaned_column_mapping.values():
                            continue  # 已经被显式映射，跳过
                        # 清理多级表头产生的列名
                        clean_col = clean_column_name(col)
                        if clean_col != col:
                            df = df.rename(columns={col: clean_col})
                            # 更新映射关系以便于记录
                            for orig, target in cleaned_column_mapping.items():
                                if target == col:
                                    cleaned_column_mapping[orig] = clean_col
                
                # 获取最终要导入的列 (重命名后的)
                final_columns = df.columns.tolist()

                # 处理数据类型 (基于 column_settings 和清理后的列名)
                sqlalchemy_dtypes = {}
                mysql_dtypes_for_history = {} # 用于记录历史
                for original_col, settings in column_settings.items():
                    # 获取映射和清理后的列名
                    final_col = cleaned_column_mapping.get(original_col)
                    if final_col and final_col in df.columns:
                        mysql_type = settings['type'].lower()
                        mysql_dtypes_for_history[final_col] = mysql_type # 记录原始请求的类型
                        base_type = mysql_type.split('(')[0]

                        # 根据基础类型进行数据转换 (同之前的逻辑)
                        if base_type in ['tinyint', 'smallint', 'mediumint', 'int', 'bigint']:
                            df[final_col] = pd.to_numeric(df[final_col], errors='coerce')
                            # 处理 NaN 值，避免转换为 float
                            df[final_col] = df[final_col].astype('Int64') # 使用 Pandas 可空整数类型
                        elif base_type in ['float', 'double', 'decimal']:
                            df[final_col] = pd.to_numeric(df[final_col], errors='coerce')
                        elif base_type in ['date', 'datetime', 'timestamp']:
                            df[final_col] = pd.to_datetime(df[final_col], errors='coerce')
                        elif base_type == 'time':
                             try:
                                df[final_col] = pd.to_datetime(df[final_col], errors='coerce').dt.time
                             except AttributeError: # 如果已经是 time 类型
                                pass
                        elif base_type == 'year':
                            try:
                                df[final_col] = pd.to_datetime(df[final_col], errors='coerce').dt.year
                                df[final_col] = df[final_col].astype('Int64') # 转为可空整数
                            except AttributeError: # 如果已经是数字类型
                                df[final_col] = pd.to_numeric(df[final_col], errors='coerce').astype('Int64')

                        elif base_type in ['boolean', 'bool', 'tinyint'] and '1' in mysql_type:
                             bool_map = {
                                 'true': True, 'false': False,
                                 '1': True, '0': False,
                                 'yes': True, 'no': False,
                                 '是': True, '否': False,
                                 '对': True, '错': False
                             }
                             # 转换为字符串处理，然后映射，最后转为 Pandas Boolean 类型
                             df[final_col] = df[final_col].astype(str).str.lower().map(bool_map).astype('boolean')
                        elif base_type == 'json':
                             df[final_col] = df[final_col].apply(lambda x: None if pd.isna(x) else str(x))
                        else: # 包括 enum, set, char, varchar, text 等都转为 string
                             df[final_col] = df[final_col].astype(str)

                        # 将 MySQL 类型转换为 SQLAlchemy 类型 (同之前的逻辑)
                        if base_type in ['tinyint', 'smallint', 'mediumint', 'int', 'bigint']:
                             sqlalchemy_dtypes[final_col] = types.BigInteger # 使用 BigInteger 以兼容 Int64
                        elif base_type in ['float', 'double', 'decimal']:
                             sqlalchemy_dtypes[final_col] = types.Float
                        elif base_type in ['date', 'datetime', 'timestamp']:
                             sqlalchemy_dtypes[final_col] = types.DateTime
                        elif base_type == 'time':
                             sqlalchemy_dtypes[final_col] = types.Time
                        elif base_type == 'year':
                             sqlalchemy_dtypes[final_col] = types.Integer
                        elif base_type in ['boolean', 'bool', 'tinyint'] and '1' in mysql_type:
                             sqlalchemy_dtypes[final_col] = types.Boolean
                        elif base_type == 'json':
                             sqlalchemy_dtypes[final_col] = types.JSON
                        elif base_type in ['char', 'varchar']:
                             length = 255
                             if '(' in mysql_type:
                                 try: length = int(mysql_type.split('(')[1].split(')')[0])
                                 except ValueError: length = 255
                             sqlalchemy_dtypes[final_col] = types.String(length=length)
                        elif base_type in ['text', 'tinytext', 'mediumtext', 'longtext']:
                             sqlalchemy_dtypes[final_col] = types.Text
                        elif base_type in ['enum', 'set']:
                             sqlalchemy_dtypes[final_col] = types.String(255) # 将 ENUM/SET 视为 String
                        else:
                             sqlalchemy_dtypes[final_col] = types.String(255) # 默认 String

                # 导入数据 (替换)
                logger.info(f"开始向表 '{table_name}' 替换数据，列: {final_columns}")
                total_rows = len(df)
                chunk_size = 1000
                df.to_sql(
                    name=table_name,
                    con=engine,
                    if_exists='replace',
                    index=False,
                    dtype=sqlalchemy_dtypes,
                    method='multi' if db_config['type'] != 'sqlite' else None,
                    chunksize=chunk_size if db_config['type'] != 'sqlite' else None
                )
                rows_imported = total_rows
                logger.info(f"成功向表 '{table_name}' 导入 {rows_imported} 行数据 (替换模式)")

                # 处理主键 (仅 MySQL 和 PostgreSQL 支持 ALTER TABLE ADD PRIMARY KEY)
                primary_key_final = cleaned_column_mapping.get(primary_key_original) if primary_key_original else None
                if primary_key_final and primary_key_final in df.columns and db_config['type'] in ['mysql', 'postgres']:
                    with engine.connect() as connection:
                        try:
                            inspector = inspect(engine)
                            pk_constraint = inspector.get_pk_constraint(table_name)
                            pk_exists = pk_constraint and primary_key_final in pk_constraint['constrained_columns']

                            if not pk_exists:
                                pk_sql = ""
                                if db_config['type'] == 'mysql':
                                     # MySQL 需要确保列存在且类型合适，这里假设类型已在 to_sql 中设置正确
                                     pk_sql = f"ALTER TABLE `{table_name}` ADD PRIMARY KEY (`{primary_key_final}`)"
                                elif db_config['type'] == 'postgres':
                                     pk_sql = f'ALTER TABLE "{table_name}" ADD PRIMARY KEY ("{primary_key_final}")'

                                if pk_sql:
                                     connection.execute(text(pk_sql))
                                     connection.commit()
                                     logger.info(f"成功为表 '{table_name}' 添加主键 '{primary_key_final}'")
                        except Exception as pk_err:
                            logger.error(f"为表 '{table_name}' 添加主键 '{primary_key_final}' 失败: {pk_err}")
                            # 不阻塞流程，但记录错误
                            pass # 可以选择是否向用户报告此错误

                elif primary_key_final and db_config['type'] == 'sqlite':
                     logger.warning(f"SQLite 不支持在表创建后添加主键。主键 '{primary_key_final}' 将不会被设置。")
                     primary_key_final = None # 标记主键未设置成功


                # 记录替换历史
                column_info_history = {
                    'mode': 'replace',
                    'columns': final_columns, # 清理和重命名后的列
                    'original_columns_selected': selected_columns, # 文件中的原始选中列
                    'column_mapping_requested': column_mapping, # 用户请求的映射关系 {original: target_request}
                    'column_mapping_cleaned': cleaned_column_mapping, # 清理后的映射 {original: final_cleaned_name}
                    'data_types_requested': {cleaned_column_mapping.get(orig): settings['type']
                                              for orig, settings in column_settings.items() if cleaned_column_mapping.get(orig)},
                    'primary_key_original': primary_key_original, # 原始选择的主键列
                    'primary_key_final': primary_key_final, # 清理和映射后的主键列 (如果设置成功)
                    'excel_structure': excel_structure if excel_structure else None
                }

            else:
                 return jsonify({"status": "error", "message": f"无效的导入模式: {import_mode}"}), 400


            # --- 通用成功处理 ---
            add_import_history(
                file_name=data.get('file_name', '未知文件'),
                file_path=data.get('file_path', '未知路径'),
                table_name=table_name,
                database_type=db_config['type'],
                total_rows=rows_imported,
                column_count=len(df.columns if import_mode == 'replace' else df_append.columns),
                column_info=column_info_history,
                status='success'
            )

            success_message = f"数据导入成功 ({'追加' if import_mode == 'append' else '替换'}模式)。共处理 {rows_imported} 行记录。"
            if import_mode == 'replace' and primary_key_final:
                 success_message += f" 主键 '{primary_key_final}' 已设置。"
            elif import_mode == 'replace' and primary_key_original and not primary_key_final:
                 success_message += f" 主键 '{primary_key_original}' 设置失败（可能由于数据库类型或列不存在）。"

            return jsonify({
                "status": "success",
                "message": success_message,
                "rows_imported": rows_imported
            })

        except Exception as e:
            logger.error(f"数据导入失败 ({import_mode}模式): {str(e)}")
            # 记录失败历史
            failed_column_info = { 
                'mode': import_mode,
                'excel_structure': excel_structure if excel_structure else None
            }
            # 尝试获取更多上下文信息
            if 'df' in locals(): failed_column_info['original_columns_selected'] = selected_columns
            if 'column_mapping' in locals(): failed_column_info['column_mapping_requested'] = column_mapping
            if 'final_mapping' in locals(): failed_column_info['final_mapping_append'] = final_mapping # append 模式
            if 'cleaned_column_mapping' in locals(): failed_column_info['column_mapping_cleaned'] = cleaned_column_mapping # replace 模式
            if 'column_settings' in locals(): failed_column_info['data_types_requested'] = column_settings
            if 'primary_key_original' in locals(): failed_column_info['primary_key_original'] = primary_key_original

            add_import_history(
                file_name=data.get('file_name', '未知文件'),
                file_path=data.get('file_path', '未知路径'),
                table_name=table_name,
                database_type=db_config.get('type', '未知'),
                total_rows=len(df) if 'df' in locals() and isinstance(df, pd.DataFrame) else 0,
                column_count=len(df.columns) if 'df' in locals() and isinstance(df, pd.DataFrame) else 0,
                column_info=failed_column_info,
                status='error',
                error_message=str(e)
            )
            # 重新抛出异常，以便外层捕获并返回500错误
            raise

        finally:
            if engine:
                engine.dispose()

    except Exception as e:
        # 这个捕获处理 engine 创建失败或者请求解析失败等早期错误
        logger.error(f"数据导入处理中发生意外错误: {str(e)}")
        return jsonify({"status": "error", "message": f"数据导入处理失败: {str(e)}"}), 500


@data_import_bp.route('/api/import-history', methods=['GET'])
def get_import_history_api():
    try:
        limit = request.args.get('limit', default=10, type=int)
        history = get_import_history(limit)
        processed_history = []
        for record in history:
            new_record = dict(record) # Create a copy
            column_info_raw = new_record.get('column_info')
            
            parsed_info = {} # Default to empty object
            if isinstance(column_info_raw, str):
                try:
                    parsed_info = json.loads(column_info_raw)
                    # Ensure the parsed result is an object (dict in Python)
                    if not isinstance(parsed_info, dict):
                         logger.warning(f"历史记录 ID {new_record.get('id')} 的 column_info 解析后不是字典: {type(parsed_info)}")
                         parsed_info = {} # Default if not dict
                except json.JSONDecodeError:
                    logger.warning(f"无法解析历史记录 ID {new_record.get('id')} 的 column_info 字符串: {column_info_raw}")
                    # Keep parsed_info as {}
            elif isinstance(column_info_raw, dict):
                 parsed_info = column_info_raw # Already a dict
            elif column_info_raw is not None:
                 # Log if it was unexpected type but not None
                 logger.warning(f"历史记录 ID {new_record.get('id')} 的 column_info 不是字符串或字典: {type(column_info_raw)}")
                 # Keep parsed_info as {}

            new_record['column_info'] = parsed_info # Assign the guaranteed object/dict
            processed_history.append(new_record)

        return jsonify({
            "status": "success",
            "history": processed_history # Return processed data
        })
    except Exception as e:
        logger.error(f"获取导入历史失败: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

# --- 新增的 API 路由 --- 
@data_import_bp.route('/api/table-data', methods=['GET'])
def get_table_data():
    """获取指定表的数据预览"""
    table_name = request.args.get('table_name')
    limit = request.args.get('limit', default=100, type=int)

    if not table_name:
        return jsonify({"status": "error", "message": "缺少表名参数"}), 400
    
    if limit <= 0:
        limit = 100

    try:
        config = vanna_manager.get_config()
        if not config or 'database' not in config:
            return jsonify({"status": "error", "message": "未配置数据库连接"}), 400

        db_config = config['database']
        engine = None
        try:
            # --- 创建数据库连接引擎 (复用之前的逻辑) --- 
            if db_config['type'] == 'mysql':
                mysql_url = f"mysql+pymysql://{db_config['username']}:{db_config['password']}@{db_config['host']}:{db_config.get('port', 3306)}/{db_config['database_name']}?charset=utf8mb4"
                engine = create_engine(mysql_url)
            elif db_config['type'] == 'postgres':
                 pg_url = f"postgresql+psycopg2://{db_config['username']}:{db_config['password']}@{db_config['host']}:{db_config.get('port', 5432)}/{db_config['database_name']}"
                 engine = create_engine(pg_url)
            elif db_config['type'] == 'sqlite':
                 sqlite_path = db_config['database_name']
                 if not sqlite_path:
                      return jsonify({"status": "error", "message": "SQLite 数据库路径未配置"}), 400
                 engine = create_engine(f'sqlite:///{sqlite_path}')
            else:
                  return jsonify({"status": "error", "message": "不支持的数据库类型"}), 400

            # --- 检查表是否存在 --- 
            inspector = inspect(engine)
            if not inspector.has_table(table_name):
                 return jsonify({"status": "error", "message": f"表 '{table_name}' 不存在"}), 404

            # --- 查询数据 --- 
            # 使用 Pandas 读取 SQL 数据，可以更好地处理数据类型
            # 需要确保表名和列名在 SQL 中正确引用
            # SQLAlchemy 的 text() 会自动处理参数绑定，但 limit 不是标准参数
            # 对于表名，我们需要手动确保安全或使用 SQLAlchemy 的更高级功能
            # 这里假设 table_name 是安全的，或者在前端已经做了基本校验
            # 注意：直接格式化 SQL 字符串可能存在注入风险，如果 table_name 来自不可信来源
            # 更好的方法是验证 table_name 只包含字母、数字、下划线等
            # if not re.match(r'^[a-zA-Z0-9_]+$', table_name):
            #     return jsonify({"status": "error", "message": "无效的表名"}), 400

            # 获取列名
            columns_info = inspector.get_columns(table_name)
            columns = [col['name'] for col in columns_info]

            # 构建查询语句
            # 使用 SQLAlchemy 的 text 和 literal_column 来安全地引用表名
            # (需要导入 literal_column from sqlalchemy.sql)
            # from sqlalchemy.sql import literal_column
            # stmt = text(f"SELECT * FROM {literal_column(table_name)} LIMIT :limit") 
            # 但 literal_column 可能不适用于所有数据库或场景，这里简化处理
            # 直接使用 f-string，依赖于表名在数据库中已存在（has_table 检查）
            # 注意：直接格式化 SQL 字符串可能存在注入风险，如果 table_name 来自不可信来源
            # 更好的方法是验证 table_name 只包含字母、数字、下划线等
            # if not re.match(r'^[a-zA-Z0-9_]+$', table_name):
            #     return jsonify({"status": "error", "message": "无效的表名"}), 400

            # 尝试用引号括起表名以兼容更多情况
            stmt = text(f'SELECT * FROM "{table_name}" LIMIT :limit') # 尝试用引号括起表名以兼容更多情况
            if db_config['type'] == 'mysql':
                 stmt = text(f'SELECT * FROM `{table_name}` LIMIT :limit')
            
            df = pd.read_sql(stmt, engine, params={'limit': limit})

            # 获取总行数
            count_stmt = text(f'SELECT COUNT(*) FROM "{table_name}"')
            if db_config['type'] == 'mysql':
                 count_stmt = text(f'SELECT COUNT(*) FROM `{table_name}`')
            total_rows = -1 # 默认为未知
            with engine.connect() as connection:
                try:
                    result = connection.execute(count_stmt)
                    total_rows = result.scalar_one_or_none()
                except SQLAlchemyError as count_err:
                     logger.warning(f"获取表 '{table_name}' 总行数失败: {count_err}")

            # --- 准备返回数据 --- 
            # 使用 pandas_to_json_safe 确保数据可序列化
            data_list = pandas_to_json_safe(df) 

            return jsonify({
                "status": "success",
                "table_name": table_name,
                "columns": columns,
                "data": data_list,
                "limit": limit,
                "total": total_rows if total_rows is not None else -1
            })

        except SQLAlchemyError as db_err:
            logger.error(f"查询表 '{table_name}' 数据时数据库错误: {db_err}")
            return jsonify({"status": "error", "message": f"数据库错误: {str(db_err)}"}), 500
        except Exception as e:
            logger.error(f"获取表 '{table_name}' 的数据失败: {str(e)}")
            return jsonify({"status": "error", "message": f"获取表数据失败: {str(e)}"}), 500
        finally:
            if engine:
                engine.dispose()

    except Exception as e:
        logger.error(f"处理获取表数据请求时发生错误: {str(e)}")
        return jsonify({"status": "error", "message": f"处理请求失败: {str(e)}"}), 500